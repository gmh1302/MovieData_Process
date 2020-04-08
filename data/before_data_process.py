from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
import requests
from openpyxl.utils.exceptions import IllegalCharacterError

url = "https://movie.naver.com/movie/bi/mi/pointWriteFormList.nhn?code=161967&type=after&onlyActualPointYn=N&onlySpoilerPointYn=N&order=newest"
html = bs(requests.get(url).content, "html.parser", from_encoding="utf-8")

# 댓글의 총 개수
cnt = html.select("body > div > div > div.score_total > strong > em")[0].contents[0].replace(',','')

############################################################################################################

# csv 파일 생성
wb = Workbook()
ws = wb.create_sheet("기생충", 0)

# csv 파일의 1번째 행을 컬럼명으로 설정
ws.cell(1,1,"아이디")
ws.cell(1,2,"날짜")
ws.cell(1,3,"평점")
ws.cell(1,4,"댓글")
ws.cell(1,5,"좋아요")
ws.cell(1,6,"싫어요")
ws.cell(1,7,"수상일 이후 평점 평균")


row = 2 # 첫 번째 행은 컬럼명이기 때문에 row라는 변수를 2를 설정

score_sum = 0 # 수상일 이후의 평점 평균 산출을 위해 네티즌의 평점의 합을 나타내는 변수를 score_sum으로 지정

for x in range(244,int(cnt)//10+2): # 244페이지부터 수상일 이전의 댓글이 존재합니다. 1부터 돌리기에는 시간이 너무 오래걸려서 융통성있게 244부터 시작했습니다.
    html  = bs(requests.get(url + "&page=" + str(x)).content, "html.parser", from_encoding="utf-8")
    size = len(html.select("body > div > div > div.score_result > ul > li > div.score_reple > p")) # 페이지당 댓글의 수만큼 for문을 돌리도록 설정(마지막 페이지에 댓글이 10개가 아닐 수도 있기 때문)

    for i in range(1,size+1):
        #유저 아이디 크롤링 코드
        name    = html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.score_reple > dl > dt > em:nth-child(1) > a > span")[0].contents[0]

        #유저의 리뷰를 단 날짜 크롤링 코드
        date    = html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.score_reple > dl > dt > em:nth-child(2)")[0].contents[0]
        date2   = int(date[0:10].replace(".", "")) #앞에 순수한 연도월일을 date2 변수에 담음.

        #유저가 부여한 평점 크롤링 코드
        score   = int(html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.star_score > em")[0].contents[0])

        # 네이버 영화 댓글 크롤링 코드
        comment_line = html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.score_reple > p")[0]
        n = int(len(comment_line))

        comment = comment_line.contents[n - 2]
        if int(len(comment)) > 1:
            comment = comment.contents[1].contents[1].contents[0].strip()
        else:
            comment = comment.contents[0].strip()

        #유저가 받은 좋아요 숫자 크롤링 코드
        like    = int(html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.btn_area > a._sympathyButton > strong")[0].contents[0])

        #유저가 받은 싫어요 숫자 크롤링 코드
        dislike = int(html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.btn_area > a._notSympathyButton > strong")[0].contents[0])

        if date2 >= 20190530 and date2 <= 20200208: # 개봉일(2019.05.30) 부터 수상일(2020.02.09) 이전의 데이터를 전처리
            # csv 파일에 전처리한 데이터를 삽입하기 위한 코드
            print(date2, '   ', score, '   ', comment)
            try: # 중간에 댓글 2~3개가 IllegalCharacterError 오류 발생이 일어나서 예외 처리 했습니다. 나중에 고쳐보도록 하겠습니다.
                ws.cell(row,1,name)
                ws.cell(row,2,date2)
                ws.cell(row,3,score)
                ws.cell(row,4,comment)
                ws.cell(row,5,like)
                ws.cell(row,6,dislike)

                score_sum+=score
                row+=1 # 한 명의 유저 데이터를 csv를 옮길 때마다 행을 1칸씩 내려주고자  변수 row에 1을 더했다.

            except IllegalCharacterError:
                pass

ws.cell(2,7,round(score_sum/(row-2),2)) # 첫 번째 행에 컬럼명 '수상일 이후의 평점 평균'을 생성하여 평점의 합에서 댓글의 수만큼 나눠서 값을 도출
wb.save("기생충_수상일 이전.xlsx")
