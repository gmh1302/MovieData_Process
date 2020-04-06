from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
import requests

url = "https://movie.naver.com/movie/bi/mi/pointWriteFormList.nhn?code=161967&type=after&onlyActualPointYn=N&onlySpoilerPointYn=N&order=newest"
html = bs(requests.get(url).content, "html.parser", from_encoding="utf-8")
cnt = html.select("body > div > div > div.score_total > strong > em")[0].contents[0].replace(',','')

############################################################################################################

# csv 파일생성
wb = Workbook()
ws = wb.create_sheet("기생충", 0)

# csv 파일의 대표 Column을 설정
ws.cell(1,1,"아이디")
ws.cell(1,2,"날짜")
ws.cell(1,3,"평점")
ws.cell(1,4,"댓글")
ws.cell(1,5,"좋아요")
ws.cell(1,6,"싫어요")
ws.cell(1,7,"평점 평균")

row = 2 # 첫 번째 row는 컬럼명이기 때문에 row를 2를 설정.

score_sum = 0

for x in range(1,int(cnt)//10+2): # 1페이지부터 수상일 이후의 댓글을 도출하고자함.
    html  = bs(requests.get(url + "&page=" + str(x)).content, "html.parser", from_encoding="utf-8")
    size = len(html.select("body > div > div > div.score_result > ul > li > div.score_reple > p")) # 페이지당 댓글의 수만큼 for문을 돌리도록 설정(마지막 페이지에 댓글이 10개가 아닐 수도 있기 때문)

    for i in range(1,size+1):
        #유저 아이디 크롤링 코드
        name    = html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.score_reple > dl > dt > em:nth-child(1) > a > span")[0].contents[0]
        #유저의 리뷰를 단 날짜 크롤링 코드
        date    = html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.score_reple > dl > dt > em:nth-child(2)")[0].contents[0]
        date2   = int(date[0:10].replace(".", ""))
        #유저가 부여한 평점 크롤링 코드
        score   = int(html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.star_score > em")[0].contents[0])

        # 네이버 영화 댓글 크롤링 코드
        comment_line = html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.score_reple > p")[0]
        n = int(len(comment_line))
        comment = comment_line.contents[n - 2]
        if int(len(comment)) > 1:
            comment = comment.contents[1].contents[1].contents[0].strip()
        else:
            comment = comment.contents[0].strip()
        #유저의 리뷰 좋아요 숫자 크롤링 코드
        like    = int(html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.btn_area > a._sympathyButton > strong")[0].contents[0])

        #유저의 리뷰 싫어요 숫자 크롤링 코드
        dislike = int(html.select("body > div > div > div.score_result > ul > li:nth-child("+str(i)+") > div.btn_area > a._notSympathyButton > strong")[0].contents[0])

        print(date2, '   ', score, '   ', comment)
        # exit()
        if date2 >= 20200210: # 수상일(2020.02.09) 이후의 데이터를 전처리
            # csv 파일에 전처리한 데이터를 삽입하기 위한 코드
            ws.cell(row,1,name)
            ws.cell(row,2,date2)
            ws.cell(row,3,score)
            ws.cell(row,4,comment)
            ws.cell(row,5,like)
            ws.cell(row,6,dislike)

            score_sum+=score

        else:
            # print("수상일 이전 네티즌 평점 : {}".format(round(score_sum/(row-2),2)))
            ws.cell(2,7,round(score_sum/(row-2),2))
            wb.save("after.xlsx")
            exit()

        # 한 명의 유저 데이터를 csv를 옮길 때마다 행을 1칸씩 내려주고자  변수 row에 1을 더했다.
        row+=1


